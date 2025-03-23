import os
import openpyxl
from flask import Flask, render_template_string, request, jsonify
import sqlite3
import random
import webbrowser
import threading

app = Flask(__name__)

EXCEL_PATH = 'torneio.xlsx'
DB_PATH = 'jogadores.db'


def abrir_navegador():
    webbrowser.open_new('http://127.0.0.1:5000')


def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS jogadores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL,
                time TEXT,
                ordem_escolha INTEGER
            )
        ''')
        conn.commit()


@app.route('/')
def home():
    return render_template_string(HOME_HTML)


@app.route('/cadastrar_jogador', methods=['POST'])
def cadastrar_jogador():
    data = request.json
    try:
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO jogadores (nome, time, ordem_escolha)
                VALUES (?, ?, ?)
            ''', (data['nome'], data.get('time'), data.get('ordem_escolha')))
            conn.commit()
        return jsonify({'message': 'Jogador cadastrado com sucesso!'})
    except Exception as e:
        return jsonify({'message': f'Erro: {str(e)}'}), 500


@app.route('/sortear_ordem_escolha', methods=['POST'])
def sortear_ordem_escolha():
    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT nome FROM jogadores')
        jogadores = cursor.fetchall()

    if len(jogadores) % 2 != 0:
        return jsonify({'message': 'Número de jogadores deve ser par.'}), 400

    # Embaralha a lista de jogadores para sortear a ordem
    random.shuffle(jogadores)
    ordem_escolha = {jogador[0]: idx + 1 for idx, jogador in enumerate(jogadores)}

    # Atualiza a ordem de escolha no banco de dados
    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.cursor()
        for nome, ordem in ordem_escolha.items():
            cursor.execute('''
                UPDATE jogadores
                SET ordem_escolha = ?
                WHERE nome = ?
            ''', (ordem, nome))
        conn.commit()

    return jsonify({'message': 'Ordem de escolha sorteada com sucesso!', 'ordem_escolha': ordem_escolha})

@app.route('/escolher_time', methods=['POST'])
def escolher_time():
    data = request.json
    nome = data['nome']
    time = data['time']

    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE jogadores
            SET time = ?
            WHERE nome = ?
        ''', (time, nome))
        conn.commit()

    return jsonify({'message': f'Time {time} escolhido por {nome} com sucesso!'})


@app.route('/sortear_torneio', methods=['POST'])
def sortear_torneio():
    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT nome, time FROM jogadores')
        jogadores = cursor.fetchall()

    if len(jogadores) % 2 != 0:
        return jsonify({'message': 'Número de jogadores deve ser par.'}), 400

    random.shuffle(jogadores)
    chave_a = jogadores[:len(jogadores)//2]
    chave_b = jogadores[len(jogadores)//2:]

    confrontos = []
    for i in range(len(chave_a)):
        confrontos.append([
            {'nome': chave_a[i][0], 'time': chave_a[i][1]},  # Jogador da chave A
            {'nome': chave_b[i][0], 'time': chave_b[i][1]}   # Jogador da chave B
        ])

    gerar_excel(confrontos)
    return jsonify({'message': 'Torneio sorteado com sucesso!', 'confrontos': confrontos})



def gerar_excel(confrontos):
    workbook = openpyxl.Workbook() if not os.path.exists(EXCEL_PATH) else openpyxl.load_workbook(EXCEL_PATH)
    sheet = workbook.active

    if sheet.max_row == 1:
        sheet.append(['Jogador 1', 'Jogador 2'])

    for confronto in confrontos:
        sheet.append([confronto[0]['nome'], confronto[1]['nome']])

    workbook.save(EXCEL_PATH)


HOME_HTML = '''
<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Sorteio de Torneio FIFA</title>
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css">
    <style>
        .hidden {
            display: none;
        }
        .resultado-container {
            margin: 20px auto;
            padding: 20px;
            background-color: #f8f9fa;
            border-radius: 10px;
            box-shadow: 0 4px 8px rgba(0, 0, 0, 0.1);
        }
        .resultado-titulo {
            font-size: 2em;
            color: #343a40;
            text-align: center;
            margin-bottom: 20px;
        }
        .resultado-tabela {
            width: 100%;
            border-collapse: collapse;
            font-size: 1.5em;
        }
        .resultado-tabela th,
        .resultado-tabela td {
            padding: 15px;
            text-align: center;
            border: 1px solid #dee2e6;
        }
        .resultado-tabela th {
            background-color: #343a40;
            color: white;
        }
        .resultado-tabela tr:nth-child(even) {
            background-color: #f2f2f2;
        }
        .resultado-tabela tr:hover {
            background-color: #ddd;
        }
        .navbar {
            margin-bottom: 20px;
        }
    </style>
</head>
<body>
    <nav class="navbar navbar-expand-lg navbar-dark bg-dark">
        <div class="container-fluid">
            <a class="navbar-brand" href="/">Torneio FIFA</a>
            <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#navbarNav" aria-controls="navbarNav" aria-expanded="false" aria-label="Toggle navigation">
                <span class="navbar-toggler-icon"></span>
            </button>
            <div class="collapse navbar-collapse" id="navbarNav">
                <ul class="navbar-nav">
                    <li class="nav-item">
                        <a class="nav-link" href="#" onclick="mostrarSecao('cadastrar')">Cadastrar Jogador</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="#" onclick="mostrarSecao('sortear-ordem')">Sortear Ordem</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="#" onclick="mostrarSecao('escolher-time')">Escolher Time</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="#" onclick="mostrarSecao('sortear-torneio')">Sortear Torneio</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="#" onclick="mostrarSecao('criador')">Criador: Willian Batista Oliveira</a>
                    </li>
                </ul>
            </div>
        </div>
    </nav>

    <main class="container my-5 text-center">
        <div id="cadastrar" class="hidden">
            <h2>Cadastrar Jogador</h2>
            <form id="formCadastro">
                <div class="mb-3">
                    <label for="nome" class="form-label">Nome</label>
                    <input type="text" class="form-control" id="nome" required>
                </div>
                <button type="submit" class="btn btn-success">Cadastrar</button>
            </form>
        </div>

        <div id="sortear-ordem" class="hidden">
            <h2>Sortear Ordem de Escolha</h2>
            <button class="btn btn-primary" onclick="sortearOrdemEscolha()">Sortear Ordem</button>
            <div id="resultado-ordem"></div>
        </div>

        <div id="escolher-time" class="hidden">
            <h2>Escolher Time</h2>
            <form id="formEscolherTime">
                <div class="mb-3">
                    <label for="nomeEscolha" class="form-label">Nome do Jogador</label>
                    <input type="text" class="form-control" id="nomeEscolha" required>
                </div>
                <div class="mb-3">
                    <label for="timeEscolha" class="form-label">Time</label>
                    <input type="text" class="form-control" id="timeEscolha" required>
                </div>
                <button type="submit" class="btn btn-warning">Escolher Time</button>
            </form>
        </div>

        <div id="sortear-torneio" class="hidden">
            <h2>Sortear Torneio</h2>
            <button class="btn btn-danger" onclick="sortearTorneio()">Sortear Torneio</button>
            <div id="resultado"></div>
        </div>

        <div id="criador" class="hidden">
            <h2>Criador do Sistema FIFA - sorteio para torneios</h2>
            <h1>Willian Batista Oliveira</h1>
            <img src="static/Curriculum 2025.jpg" alt="Imagem do Criador" style="width: 550px; height: 1000px;">
        </div>
    </main>

    <script>
        // Função para mostrar a seção clicada e esconder as outras
        function mostrarSecao(secao) {
            document.querySelectorAll('main > div').forEach(div => {
                div.classList.add('hidden');
            });
            document.getElementById(secao).classList.remove('hidden');
        }

        document.getElementById('formCadastro').addEventListener('submit', async (event) => {
            event.preventDefault();
            const nome = document.getElementById('nome').value;

            try {
                const response = await fetch('/cadastrar_jogador', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ nome })
                });

                const data = await response.json();
                alert(data.message);
                document.getElementById('formCadastro').reset();
            } catch (error) {
                console.error('Erro ao cadastrar:', error);
            }
        });

        async function sortearOrdemEscolha() {
            try {
                const response = await fetch('/sortear_ordem_escolha', { method: 'POST' });
                const data = await response.json();
                alert(data.message);
                exibirOrdemEscolha(data.ordem_escolha);
            } catch (error) {
                console.error('Erro ao sortear ordem:', error);
            }
        }

        function exibirOrdemEscolha(ordem_escolha) {
            const resultadoDiv = document.getElementById('resultado-ordem');
            resultadoDiv.innerHTML = '';

            // Cria a tabela
            const tabela = document.createElement('table');
            tabela.className = 'resultado-tabela';

            // Cria o cabeçalho da tabela
            const cabecalho = document.createElement('thead');
            const linhaCabecalho = document.createElement('tr');
            const colunaNome = document.createElement('th');
            colunaNome.textContent = 'Nome';
            const colunaOrdem = document.createElement('th');
            colunaOrdem.textContent = 'Ordem';
            linhaCabecalho.appendChild(colunaNome);
            linhaCabecalho.appendChild(colunaOrdem);
            cabecalho.appendChild(linhaCabecalho);
            tabela.appendChild(cabecalho);

            // Cria o corpo da tabela
            const corpoTabela = document.createElement('tbody');

            // Adiciona cada jogador e sua ordem na tabela
            for (const [nome, ordem] of Object.entries(ordem_escolha)) {
                const linha = document.createElement('tr');
                const celulaNome = document.createElement('td');
                celulaNome.textContent = nome;
                const celulaOrdem = document.createElement('td');
                celulaOrdem.textContent = ordem;
                linha.appendChild(celulaNome);
                linha.appendChild(celulaOrdem);
                corpoTabela.appendChild(linha);
            }

            tabela.appendChild(corpoTabela);
            resultadoDiv.appendChild(tabela);
        }

        document.getElementById('formEscolherTime').addEventListener('submit', async (event) => {
            event.preventDefault();
            const nome = document.getElementById('nomeEscolha').value;
            const time = document.getElementById('timeEscolha').value;

            try {
                const response = await fetch('/escolher_time', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ nome, time })
                });

                const data = await response.json();
                alert(data.message);
                document.getElementById('formEscolherTime').reset();
            } catch (error) {
                console.error('Erro ao escolher time:', error);
            }
        });

        async function sortearTorneio() {
            try {
                const response = await fetch('/sortear_torneio', { method: 'POST' });
                const data = await response.json();
                alert(data.message);
                exibirConfrontos(data.confrontos);
            } catch (error) {
                console.error('Erro ao realizar o sorteio:', error);
            }
        }

        function exibirConfrontos(confrontos) {
            const resultadoDiv = document.getElementById('resultado');
            resultadoDiv.innerHTML = '';

            // Cria o container do resultado
            const container = document.createElement('div');
            container.className = 'resultado-container';

            // Adiciona o título
            const titulo = document.createElement('h2');
            titulo.className = 'resultado-titulo';
            titulo.textContent = 'Confrontos do Torneio';
            container.appendChild(titulo);

            // Cria a tabela
            const tabela = document.createElement('table');
            tabela.className = 'resultado-tabela';

            // Cria o cabeçalho da tabela
            const cabecalho = document.createElement('thead');
            const linhaCabecalho = document.createElement('tr');
            const colunaCasa = document.createElement('th');
            colunaCasa.textContent = 'Casa';
            const colunaNumero = document.createElement('th');
            colunaNumero.textContent = 'Número da Partida';
            const colunaVisitante = document.createElement('th');
            colunaVisitante.textContent = 'Visitante';
            linhaCabecalho.appendChild(colunaCasa);
            linhaCabecalho.appendChild(colunaNumero);
            linhaCabecalho.appendChild(colunaVisitante);
            cabecalho.appendChild(linhaCabecalho);
            tabela.appendChild(cabecalho);

            // Cria o corpo da tabela
            const corpoTabela = document.createElement('tbody');

            // Adiciona cada confronto na tabela
            confrontos.forEach((confronto, index) => {
                const linha = document.createElement('tr');

                // Coluna Casa: Jogador da chave A e seu time
                const celulaCasa = document.createElement('td');
                celulaCasa.textContent = `${confronto[0].nome} (${confronto[0].time})`;

                // Coluna Número da Partida: Número sequencial
                const celulaNumero = document.createElement('td');
                celulaNumero.textContent = index + 1;

                // Coluna Visitante: Jogador da chave B e seu time
                const celulaVisitante = document.createElement('td');
                celulaVisitante.textContent = `${confronto[1].nome} (${confronto[1].time})`;

                // Adiciona as células à linha
                linha.appendChild(celulaCasa);
                linha.appendChild(celulaNumero);
                linha.appendChild(celulaVisitante);

                // Adiciona a linha ao corpo da tabela
                corpoTabela.appendChild(linha);
            });

            tabela.appendChild(corpoTabela);
            container.appendChild(tabela);
            resultadoDiv.appendChild(container);
        }
    </script>
</body>
</html>
'''

if __name__ == '__main__':
    init_db()
    threading.Timer(1.5, abrir_navegador).start()
    app.run(debug=True)